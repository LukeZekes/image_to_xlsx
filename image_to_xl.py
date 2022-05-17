import argparse
import ntpath
import sys
from pathvalidate.argparse import ValidationError, validate_filepath, validate_filepath_arg
from PIL import Image, UnidentifiedImageError
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter

def to2ByteHex(x):
  return hex(x).lstrip('0x').zfill(2)

def positiveInt(x):
  if x.isdigit():
    if int(x) <= 256:
      return int(x)
  raise argparse.ArgumentTypeError("-n must be a positive integer <= 256")

def validateOutputPath(x):
  try:
    validate_filepath(x)
    ext = str(ntpath.splitext(x)[1])
    if (ext == ".xlsx"):
      return x
    raise Exception
  except ValidationError as e:
    print("Invalid output path provided: " + e)
    raise argparse.ArgumentTypeError("-o must be a valid path to a .xlsx file")
  except Exception:
    raise argparse.ArgumentTypeError("-o must be a valid path to a .xlsx file")


def img_to_xlsx(image_paths = [], num_colors = 2, out_path = "output.png") :
  wb = Workbook()
  sheet = wb.active
  wb.remove(sheet)
  cell_height = 15
  cell_width = cell_height / 5.5
  for file_path in image_paths:
    file_name = ntpath.basename(file_path)
    print('Processing ' + file_name + '...')

    print('Quantizing image...')

    image = Image.open(file_path)
    quantized_image = image.quantize(num_colors).convert("RGB")
    w, h = quantized_image.size
    pixels = list(quantized_image.getdata())
    
    # Adjust the column sizes so that the cells are square
    ws = wb.create_sheet(file_name)
    print('Sizing worksheet cells...')
    for i in range (1, w + 1):
      idx = get_column_letter(i)
      ws.column_dimensions[idx].width = cell_width
    for i in range (1, h + 1):
      ws.row_dimensions[i].height = cell_height

    # Loop through every cell in a HxW grid, where H represents the height of the image and W represents its width
    print('Coloring worksheet...')
    for i in range(quantized_image.height):
      for j in range(quantized_image.width):
        c = ws.cell(i + 1, j + 1)
        # Extract the colors from the quantized image into 3 separate components
        mapToRGB = map(to2ByteHex, pixels[(i * quantized_image.width) + j])
        colors = tuple(mapToRGB)
        rgb_val = ''.join(colors)
        c.fill = PatternFill("solid", start_color = rgb_val)

  print('Saving to ' + out_path)
  wb.save(out_path)
  print('Done!')
  return out_path

def parse_args(cmd_args):
  parser = argparse.ArgumentParser(exit_on_error=False)
  parser.add_argument("-i", "--image_paths", nargs = "+", required = True, help = "the file path(s) of the original image file", type = validate_filepath_arg)
  parser.add_argument("-n", required = True, dest = "num_colors", help = "the number of colors in the spreasheet image", type = positiveInt)
  parser.add_argument("-o", metavar = "output_path", type = validateOutputPath, default = "output.xlsx", dest = "out_path", help = "the file path the created spreadsheet will be output to, defaults to output.xlsx")
  
  try:
    args = parser.parse_args(cmd_args)
  except argparse.ArgumentError as e:
    raise e
  return args

if __name__ == "__main__":
  args = parse_args(sys.argv[1:])
  img_to_xlsx(args.image_paths, args.num_colors, args.out_path)